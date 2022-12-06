import openpyxl
# Charger le workbook
file = "C:\\Users\\User\\OneDrive\\Bureau\\Testauto\\userData.xlsx"
wk = openpyxl.load_workbook(file)
sh = wk['data']
rows = sh.max_row   # Pour definir le nombre de lignes qui ont des donnees
cols = sh.max_column  # Pour definir le nombre de colonnes qui ont des donnees

# Afficher les donnees des fichiers
# 2 : c'est le pas
# for i in range (1,10,2):
#     print(i)

for r in range(1, rows + 1):
    for c in range(1, cols +1):
        print(sh.cell(r,c).value, end="   ")
    print()


